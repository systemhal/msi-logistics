import openpyxl
import os
import datetime
from config import SPREADSHEETS

class DataManager:
    def __init__ (self):
        self._cache = {}
        self._mtime = {}

    def _get_workbook(self, sp_id):
        sp_meta = SPREADSHEETS.get(sp_id)
        if not sp_meta:
            raise ValueError(f"Spreadsheet '{sp_id}' no encontrado.")
        file_path = sp_meta["path"]
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"Archivo no encontrado: {file_path}")
        
        mtime = os.path.getmtime(file_path)
        if sp_id not in self._cache or self._mtime.get(sp_id) != mtime:
            wb = openpyxl.load_workbook(file_path, data_only=True)
            self._cache[sp_id] = wb
            self._mtime[sp_id] = mtime
            
        return self._cache[sp_id], file_path

    def get_sheet_data(self, sp_id, sheet_name, page=1, page_size=25, search="", sort_by=None, sort_dir="asc"):
        sp_meta = SPREADSHEETS.get(sp_id)
        if not sp_meta or sheet_name not in sp_meta["sheets"]:
            return {"rows": [], "columns": [], "totalRows": 0, "page": 1, "totalPages": 0, "pageSize": page_size}

        header_row = 1 if sheet_name == 'Seguro de carga internacional' else 4
        data_start_row = header_row + 1

        file_path = sp_meta["path"]
        wb = openpyxl.load_workbook(file_path, data_only=True)
        if sheet_name not in wb.sheetnames:
            return {"rows": [], "columns": [], "totalRows": 0, "page": 1, "totalPages": 0, "pageSize": page_size}

        ws = wb[sheet_name]
        max_r = ws.max_row
        max_c = ws.max_column

        # Dynamic clean columns extraction
        raw_headers = [ws.cell(row=header_row, column=c).value for c in range(1, max_c + 1)]
        cols_config = []

        for c_idx, h_val in enumerate(raw_headers, start=1):
            label = str(h_val).strip() if h_val is not None else ""
            if not label:
                continue

            l_upper = label.upper()

            # Filter out dummy columns ("Columna 1", "Column 20", etc.)
            if l_upper.startswith("COLUMNA") or l_upper.startswith("COLUMN"):
                continue

            # In Registro Carga, filter out Marca temporal, Column 20, Column 21
            if sheet_name == 'Registro Carga':
                if "MARCA TEMPORAL" in l_upper or l_upper in ["COLUMN 20", "COLUMN 21"]:
                    continue

            is_auto = ("MARCA TEMPORAL" in l_upper or l_upper in ["ID", "CODIGO_CLIENTE"])
            is_select = False

            if "DIRECCION" not in l_upper:
                is_select = any(k in l_upper for k in ["SERVICIO", "TIPO", "ESTADO", "INCOTERM", "CARGA LCL", "COURRIER", "CANAL"])

            cols_config.append({
                "col": c_idx,
                "key": f"col_{c_idx}",
                "label": label,
                "auto": is_auto,
                "editable": not is_auto,
                "type": "select" if is_select else "text"
            })

        all_rows = []
        search_lower = search.strip().lower() if search else ""

        for r in range(data_start_row, max_r + 1):
            row_dict = {"_row_num": r}
            is_empty = True
            search_str = ""

            # Detectar color de fondo pintado en la celda en openpyxl
            row_bg = None
            for c_idx in range(1, min(max_c + 1, 10)):
                cell = ws.cell(row=r, column=c_idx)
                if cell.fill and cell.fill.start_color and cell.fill.start_color.rgb:
                    rgb = str(cell.fill.start_color.rgb)
                    if rgb and str(rgb) not in ['00000000', 'FFFFFFFF', '00FFFFFF', 'None']:
                        hex_val = '#' + (rgb[2:] if len(rgb) == 8 else rgb).lower()
                        if hex_val not in ['#ffffff', '#fff']:
                            row_bg = hex_val
                            break

            if row_bg:
                if any(g in row_bg for g in ['c6e0b4', 'dcfce7', 'e2efda', 'a8d08d']):
                    row_dict['_sheet_color'] = 'VERDE'
                elif any(g in row_bg for g in ['d9d9d9', 'f1f5f9', 'e7e6e6', 'bfbfbf']):
                    row_dict['_sheet_color'] = 'GRIS'
                elif any(g in row_bg for g in ['fcf7e8', 'fef9c3', 'fff2cc', 'ffe699']):
                    row_dict['_sheet_color'] = 'CREMA'
                elif any(g in row_bg for g in ['fee2e2', 'fce4d6', 'f8cbad']):
                    row_dict['_sheet_color'] = 'ROJO'
                else:
                    row_dict['_sheet_color'] = 'CUSTOM'
                    row_dict['_custom_hex'] = row_bg

            for col_meta in cols_config:
                c_idx = col_meta["col"]
                col_key = col_meta["key"]
                cell_val = ws.cell(row=r, column=c_idx).value

                if cell_val is not None and str(cell_val).strip() != "":
                    is_empty = False

                formatted_val = self._format_val(cell_val, col_meta.get("type"))
                row_dict[col_key] = formatted_val
                search_str += " " + str(formatted_val).lower()

            if not is_empty:
                if not search_lower or search_lower in search_str:
                    all_rows.append(row_dict)

        if sort_by and sort_by in [c["key"] for c in cols_config]:
            reverse = (sort_dir.lower() == "desc")
            all_rows.sort(key=lambda x: str(x.get(sort_by, "")).lower(), reverse=reverse)

        total_rows = len(all_rows)
        total_pages = max(1, (total_rows + page_size - 1) // page_size)
        page = max(1, min(page, total_pages))

        start_idx = (page - 1) * page_size
        end_idx = start_idx + page_size
        paginated_rows = all_rows[start_idx:end_idx]

        return {
            "rows": paginated_rows,
            "columns": cols_config,
            "totalRows": total_rows,
            "page": page,
            "totalPages": total_pages,
            "pageSize": page_size
        }

    def update_cell(self, sp_id, sheet_name, row_num, col_key, new_value):
        sp_meta = SPREADSHEETS.get(sp_id)
        if not sp_meta or sheet_name not in sp_meta["sheets"]:
            raise ValueError("Hoja no encontrada.")

        col_idx = int(col_key.replace("col_", ""))
        file_path = sp_meta["path"]

        wb = openpyxl.load_workbook(file_path)
        ws = wb[sheet_name]

        ws.cell(row=row_num, column=col_idx).value = new_value
        wb.save(file_path)

        if sp_id in self._mtime:
            self._mtime[sp_id] = os.path.getmtime(file_path)

        return {"success": True, "message": "Celda actualizada correctamente."}

    def update_row(self, sp_id, sheet_name, row_num, row_data):
        sp_meta = SPREADSHEETS.get(sp_id)
        if not sp_meta or sheet_name not in sp_meta["sheets"]:
            raise ValueError("Hoja no encontrada.")

        file_path = sp_meta["path"]
        wb = openpyxl.load_workbook(file_path)
        ws = wb[sheet_name]

        for k, v in row_data.items():
            if k.startswith("col_"):
                col_idx = int(k.replace("col_", ""))
                ws.cell(row=row_num, column=col_idx).value = v

        wb.save(file_path)

        if sp_id in self._mtime:
            self._mtime[sp_id] = os.path.getmtime(file_path)

        return {"success": True, "message": "Fila actualizada correctamente."}

    def add_row(self, sp_id, sheet_name, row_data):
        sp_meta = SPREADSHEETS.get(sp_id)
        if not sp_meta or sheet_name not in sp_meta["sheets"]:
            raise ValueError("Hoja no encontrada.")

        file_path = sp_meta["path"]
        wb = openpyxl.load_workbook(file_path)
        ws = wb[sheet_name]

        new_row_idx = ws.max_row + 1
        now_str = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        max_c = ws.max_column or 10
        ws.cell(row=new_row_idx, column=1).value = now_str

        for k, v in row_data.items():
            if k.startswith("col_"):
                c_idx = int(k.replace("col_", ""))
                if c_idx > 1 and c_idx <= max_c:
                    ws.cell(row=new_row_idx, column=c_idx).value = v

        wb.save(file_path)
        if sp_id in self._mtime:
            self._mtime[sp_id] = os.path.getmtime(file_path)

        return {"success": True, "message": "Registro agregado correctamente.", "row_num": new_row_idx}

    def delete_row(self, sp_id, sheet_name, row_num):
        sp_meta = SPREADSHEETS.get(sp_id)
        if not sp_meta or sheet_name not in sp_meta["sheets"]:
            raise ValueError("Hoja no encontrada.")

        file_path = sp_meta["path"]
        wb = openpyxl.load_workbook(file_path)
        ws = wb[sheet_name]

        ws.delete_rows(row_num)
        wb.save(file_path)

        if sp_id in self._mtime:
            self._mtime[sp_id] = os.path.getmtime(file_path)

        return {"success": True, "message": "Registro eliminado correctamente."}

    def get_all_stats(self):
        stats = {}
        for sp_id, sp_meta in SPREADSHEETS.items():
            stats[sp_id] = {}
            file_path = sp_meta["path"]
            if not os.path.exists(file_path):
                continue
            wb = openpyxl.load_workbook(file_path, data_only=True)
            for sname in sp_meta["sheets"]:
                if sname in wb.sheetnames:
                    ws = wb[sname]
                    start = 1 if sname == 'Seguro de carga internacional' else 4
                    cnt = max(0, ws.max_row - start)
                    stats[sp_id][sname] = cnt
        return stats

    def _format_val(self, val, col_type):
        if val is None:
            return ""
        if isinstance(val, (datetime.datetime, datetime.date)):
            if col_type == "datetime":
                return val.strftime("%d/%m/%Y %H:%M")
            return val.strftime("%d/%m/%Y")
        return str(val)

db = DataManager()
